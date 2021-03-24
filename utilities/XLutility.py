import openpyxl
def getrowcount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return(sheet.max_row)

def getcolumncount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return(sheet.max_column)

def readdata(file,sheetname,rowno,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rowno,column=columnno).value

def writedata(file,sheetname,rowno,columnno,Data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rowno,column=columnno).value=Data
    workbook.save(file)



